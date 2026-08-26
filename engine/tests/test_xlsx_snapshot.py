import sqlite3
from pathlib import Path

import openpyxl

from scientific_reading.library_service import LibraryService
from scientific_reading.models import PaperMetadata
from scientific_reading.xlsx_snapshot import XlsxSnapshotService, XLSX_COLUMNS


def _seed(root: Path, count: int = 2) -> None:
    service = LibraryService(root)
    for i in range(count):
        service.ingest(PaperMetadata(title=f"中文文献 {i}", authors=[f"作者{i}"], year=2024 + i, journal="期刊", doi=f"10.1000/{i}", pmid=str(100+i), abstract_en="English", abstract_zh="中文摘要"))
    service.close()


def test_snapshot_has_fixed_columns_all_rows_and_readme_sheet(tmp_path):
    _seed(tmp_path, 3)
    result = XlsxSnapshotService(tmp_path).refresh()
    assert result["status"] == "success"
    workbook = openpyxl.load_workbook(tmp_path / "library" / "scientific-reading.xlsx", read_only=True)
    assert workbook.sheetnames == ["文献", "说明"]
    sheet = workbook["文献"]
    assert tuple(cell.value for cell in next(sheet.iter_rows())) == XLSX_COLUMNS
    rows = list(sheet.iter_rows(values_only=True))
    assert len(rows) == 4
    assert rows[1][0] == "中文文献 0"
    assert "个人想法" not in [cell.value for row in workbook["说明"].iter_rows() for cell in row]
    workbook.close()


def test_snapshot_preserves_source_url_from_sqlite(tmp_path):
    source_url = "https://publisher.example/papers/42"
    service = LibraryService(tmp_path)
    service.ingest(
        PaperMetadata(title="URL paper", doi="10.1000/url", source_url=source_url)
    )
    service.close()

    assert XlsxSnapshotService(tmp_path).refresh()["status"] == "success"
    workbook = openpyxl.load_workbook(
        tmp_path / "library" / "scientific-reading.xlsx", read_only=True
    )
    rows = list(workbook["文献"].iter_rows(values_only=True))
    workbook.close()
    assert rows[1][12] == source_url


def test_permission_error_keeps_old_file_and_records_pending_then_retry(tmp_path, monkeypatch):
    _seed(tmp_path, 1)
    service = XlsxSnapshotService(tmp_path)
    service.refresh()
    target = tmp_path / "library" / "scientific-reading.xlsx"
    old = target.read_bytes()
    real_replace = __import__("os").replace
    def locked(src, dst):
        if str(dst) == str(target):
            raise PermissionError("locked")
        return real_replace(src, dst)
    monkeypatch.setattr("scientific_reading.xlsx_snapshot.os.replace", locked)
    result = service.refresh()
    assert result["status"] == "pending"
    assert result["error"]["code"] == "xlsx_replace_permission_denied"
    assert target.read_bytes() == old
    conn = sqlite3.connect(tmp_path / "library.sqlite")
    meta = dict(conn.execute("SELECT key,value FROM library_meta"))
    conn.close()
    assert meta["xlsx_pending"] == "1"
    assert meta["xlsx_error"] == "xlsx_replace_permission_denied"
    monkeypatch.setattr("scientific_reading.xlsx_snapshot.os.replace", real_replace)
    assert service.refresh()["status"] == "success"
    conn = sqlite3.connect(tmp_path / "library.sqlite")
    meta = dict(conn.execute("SELECT key,value FROM library_meta"))
    conn.close()
    assert meta.get("xlsx_pending") == "0"
    assert meta.get("xlsx_error") in (None, "")
